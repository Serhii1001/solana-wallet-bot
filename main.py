import os
import requests
import telebot
from flask import Flask, request
from datetime import datetime
from openpyxl import Workbook

# Configuration
TELEGRAM_TOKEN = os.getenv("BOT_TOKEN")
HELIUS_API_KEY   = os.getenv("HELIUS_API_KEY")
WEBHOOK_URL      = os.getenv("WEBHOOK_URL")  # e.g. https://your-app.onrender.com
DEXSCREENER_BASE = "https://api.dexscreener.com/latest/dex/tokens/solana/"
SOL_PRICE        = os.getenv("SOL_PRICE", "0")

# Initialize bot and app
bot = telebot.TeleBot(TELEGRAM_TOKEN)
app = Flask(__name__)

# Configure webhook (remove existing, then set new)
bot.remove_webhook()
bot.set_webhook(f"{WEBHOOK_URL}/{TELEGRAM_TOKEN}")

# Health-check endpoint
def health_check():
    return "OK", 200
app.add_url_rule('/', 'health_check', health_check, methods=['GET'])

# Telegram webhook endpoint
def telegram_webhook():
    data = request.get_data(as_text=True)
    update = telebot.types.Update.de_json(data)
    bot.process_new_updates([update])
    return "OK", 200
app.add_url_rule(f'/{TELEGRAM_TOKEN}', 'telegram_webhook', telegram_webhook, methods=['POST'])

# HTTP helper
def safe_request(url, params=None):
    for _ in range(3):
        try:
            r = requests.get(url, params=params, timeout=10)
            if r.status_code == 200:
                return r.json()
        except:
            continue
    return {}

# Helpers

def get_symbol(mint):
    return safe_request(f"https://api.helius.xyz/v0/mints/{mint}?api-key={HELIUS_API_KEY}").get('symbol', mint)

def get_historical_mcap(mint, ts):
    chart = safe_request(f"{DEXSCREENER_BASE}{mint}/chart?interval=1h").get('chart', [])
    if not chart:
        return ''
    target = int(ts.timestamp() * 1000)
    best = min(chart, key=lambda p: abs(p.get('timestamp', 0) - target))
    return best.get('marketCap', '')

def get_current_mcap(mint):
    return safe_request(f"{DEXSCREENER_BASE}{mint}").get('stats', {}).get('marketCap', '')

def format_duration(start, end):
    if not start or not end:
        return '-'
    delta = end - start
    days, rem = divmod(delta.total_seconds(), 86400)
    hours, rem = divmod(rem, 3600)
    minutes, seconds = divmod(rem, 60)
    if days:
        return f"{int(days)}d {int(hours)}h"
    if hours:
        return f"{int(hours)}h {int(minutes)}m"
    if minutes:
        return f"{int(minutes)}m"
    return f"{int(seconds)}s"

# Core logic

DEBUG_CHAT_ID = 769361377  # <-- замени на свой Telegram chat_id  # ЗАМЕНИ на свой chat_id

def welcome(m):
    bot.reply_to(m, f"Твой chat_id: {m.chat.id}")

def debug(msg):
    try:
        bot.send_message(DEBUG_CHAT_ID, f"🪵 {msg}")
    except:
        pass

def analyze_wallet(wallet):
    debug(f"Анализируем кошелёк: {wallet}")
    url = f"https://api.helius.xyz/v0/addresses/{wallet}/transactions?api-key={HELIUS_API_KEY}&limit=100"
    txs = safe_request(url) or []
    debug(f"Найдено транзакций: {len(txs)}")

    bal = safe_request(f"https://api.helius.xyz/v0/addresses/{wallet}/balances?api-key={HELIUS_API_KEY}")
    balance = bal.get('nativeBalance', 0) / 1e9
    tokens = {}
    added_tokens = 0

    for tx in txs:
        ts = datetime.fromtimestamp(tx.get('timestamp', 0))
        sig = tx.get('signature', 'unknown')
        debug(f"📦 Транзакция {sig} @ {ts}")
        sol_change = sum(n.get('amount', 0) for n in tx.get('nativeTransfers', []) if n.get('fromUserAccount') == wallet) / 1e9

        token_transfers = tx.get('tokenTransfers', [])
        for tr in token_transfers:
            mint = tr.get('mint')
            if not mint:
                continue
            debug(f"🎯 Mint найден: {mint}")

            amt = float(tr.get('tokenAmount', {}).get('uiAmount', 0))
            if amt == 0:
                debug(f"⚠️ Пропущено: amount == 0")
                continue

            owner = tr.get('tokenAmount', {}).get('owner', '')
            direction = None
            if wallet == owner and tr.get('toUserAccount') == owner:
                direction = 'buy'
            elif wallet == owner and tr.get('fromUserAccount') == owner:
                direction = 'sell'

            if direction is None:
                debug(f"⚠️ Пропущено: не определено направление (from={tr.get('fromUserAccount')}, to={tr.get('toUserAccount')})")
                continue

            rec = tokens.setdefault(mint, {
                'mint': mint,
                'symbol': get_symbol(mint),
                'spent_sol': 0,
                'earned_sol': 0,
                'buys': 0,
                'sells': 0,
                'in_tokens': 0,
                'out_tokens': 0,
                'fee': 0,
                'first_ts': None,
                'last_ts': None,
                'first_mcap': '',
                'last_mcap': '',
                'current_mcap': ''
            })

            if direction == 'buy':
                rec['buys'] += 1
                rec['in_tokens'] += amt
                rec['spent_sol'] += sol_change
                if not rec['first_ts']:
                    rec['first_ts'] = ts
                    rec['first_mcap'] = get_historical_mcap(mint, ts)
            else:
                rec['sells'] += 1
                rec['out_tokens'] += amt
                rec['earned_sol'] += sol_change
                rec['last_ts'] = ts
                rec['last_mcap'] = get_historical_mcap(mint, ts)

            rec['fee'] += tx.get('fee', 0) / 1e9
            debug(f"✅ {direction}: {amt} токенов, {sol_change:.4f} SOL")
            added_tokens += 1

    for rec in tokens.values():
        rec['delta_sol'] = rec['earned_sol'] - rec['spent_sol']
        rec['delta_pct'] = (rec['delta_sol'] / rec['spent_sol'] * 100) if rec['spent_sol'] else 0
        rec['period'] = format_duration(rec['first_ts'], rec['last_ts'])
        rec['last_trade'] = rec['last_ts'] or rec['first_ts']
        rec['current_mcap'] = get_current_mcap(rec['mint'])

    summary = {
        'wallet': wallet,
        'balance': balance,
        'pnl': sum(r['delta_sol'] for r in tokens.values()),
        'avg_win_pct': sum(r['delta_pct'] for r in tokens.values() if r['delta_sol'] > 0) / max(1, sum(1 for r in tokens.values() if r['delta_sol'] > 0)),
        'pnl_loss': sum(r['delta_sol'] for r in tokens.values() if r['delta_sol'] < 0),
        'balance_change': (sum(r['delta_sol'] for r in tokens.values()) / ((balance - sum(r['delta_sol'] for r in tokens.values())) or 1) * 100),
        'winrate': sum(1 for r in tokens.values() if r['delta_sol'] > 0) / max(1, sum(1 for r in tokens.values() if abs(r['delta_sol']) > 0)) * 100,
        'time_period': '30 days',
        'sol_price': SOL_PRICE
    }

    debug(f"📈 Всего добавлено токенов в отчёт: {added_tokens}")
    return tokens, summary



# Excel report

def generate_excel(wallet, tokens, summary):
    fn = f"{wallet}_report.xlsx"; wb = Workbook(); ws = wb.active; ws.title = "ArGhost table"
    hdr = ['Wallet','WinRate','PnL R','Avg Win %','PnL Loss','Balance change','TimePeriod','SOL Price Now','Balance']
    for i,t in enumerate(hdr,1): ws.cell(1,i,t)
    vals = [wallet, f"{summary['winrate']:.2f}%", f"{summary['pnl']:.2f} SOL", f"{summary['avg_win_pct']:.2f}%", f"{summary['pnl_loss']:.2f} SOL", f"{summary['balance_change']:.2f}%", summary['time_period'], f"{summary['sol_price']} $", f"{summary['balance']:.2f} SOL"]
    for i,v in enumerate(vals,1): ws.cell(2,i,v)
    ws.cell(4,1,'Tokens entry MCAP:'); ranges=['<5k','5k-30k','30k-100k','100k-300k','300k+']
    for i,r in enumerate(ranges,2): ws.cell(5,i,r)
    cols = ['Token','Spent SOL','Earned SOL','Delta Sol','Delta %','Buys','Sells','Last trade','Income','Outcome','Fee','Period','First buy Mcap','Last tx Mcap','Current Mcap','Contract','Dexscreener','Photon']
    for i,c in enumerate(cols,1): ws.cell(8,i,c)
    r=9
    for rec in tokens.values():
        ws.cell(r,1,rec['symbol']); ws.cell(r,2,f"{rec['spent_sol']:.2f} SOL"); ws.cell(r,3,f"{rec['earned_sol']:.2f} SOL"); ws.cell(r,4,f"{rec['delta_sol']:.2f}"); ws.cell(r,5,f"{rec['delta_pct']:.2f}%"); ws.cell(r,6,rec['buys']); ws.cell(r,7,rec['sells']);
        if rec['last_trade']: ws.cell(r,8,rec['last_trade'].strftime('%d.%m.%Y'))
        ws.cell(r,9,rec['in_tokens']); ws.cell(r,10,rec['out_tokens']); ws.cell(r,11,f"{rec['fee']:.2f}"); ws.cell(r,12,rec['period']); ws.cell(r,13,rec['first_mcap']); ws.cell(r,14,rec['last_mcap']); ws.cell(r,15,rec['current_mcap']); ws.cell(r,16,rec['mint']);
        d=ws.cell(r,17); d.value='View trades'; d.hyperlink=f"https://dexscreener.com/solana/{rec['mint']}?maker={wallet}"; p=ws.cell(r,18); p.value='View trades'; p.hyperlink=f"https://photon-sol.tinyastro.io/en/lp/{rec['mint']}"
        r+=1
    wb.save(fn); return fn

# Handlers

bot.register_message_handler(welcome, commands=['start'])

def handle(m): wallet=m.text.strip(); bot.reply_to(m,"Обрабатываю..."); tokens,summary=analyze_wallet(wallet); f=generate_excel(wallet,tokens,summary); bot.send_document(m.chat.id, open(f,'rb'))
bot.register_message_handler(handle, func=lambda _: True)

# Run app
def main(): app.run(host='0.0.0.0', port=int(os.environ.get('PORT',5000)))
if __name__=='__main__': main()
